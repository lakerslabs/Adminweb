# -*- coding: utf-8 -*-
"""
Created on Mon Jun  1 11:44:50 2020

@author: eduardo.berga
"""


from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select

from selenium.webdriver.support import expected_conditions as EC

import win32com.client as win32
import win32print
from selenium import webdriver
import time
import openpyxl
import math
import autoit

# import SelectPrint as sp
# import tkinter as tk

# #>>> Modulo para seleccionar Impresora <<<
# root = tk.Tk()
# sp.PrinterManager(root)2
# root.mainloop()
# #>>>>>>>>>>>*<<<<<<<<<<<<<

# workbook  = openpyxl.load_workbook(r'C:\Users\eduardo.berga\Downloads\EjemploImportacionMasivaEncomiendaDomicilio.xlsx')
# sheet = workbook.active
# i=1
# encabezados=[]
# while not(sheet.cell(row=2, column=i).value == None):
#     encabezados.append(sheet.cell(row=2, column=i).value)
#     i= i+1

t_0 = time.time() # Grabar el tiempo de inicio
regMaximos = 50

################################################
#       Cambio de impresora por rotuladora
################################################

# tempprinter = "Microsoft Print to PDF"
tempprinter = "\\\\XL-FPS\HP Administracion"
#"\\\\xl-pc-fac06\\ZDesigner GC420t (EPL)"
currentprinter = win32print.GetDefaultPrinter()
win32print.SetDefaultPrinter(tempprinter)
print('Se establecio la siguiente impresora de rotulos: \n'+ tempprinter)

################################################
################################################

workbook  = openpyxl.load_workbook(r'X:\Logistica\1. Logistica Interno\Historiales_de_Logistica\ImportacionMasivaEncomiendaDomicilio.xlsx')
sheet = workbook.active

listaNom=[]
NomCli = str()
codDirec=0
i=3
while not(sheet.cell(row=i, column=1).value == None):
    aux=sheet.cell(row=i, column=7).value
    listaNom=aux.split(" ")
    aux01=len(listaNom)
    apellido=listaNom[aux01-1]
    h=0
    NomCli=''
    while h < aux01-1:
        NomCli= NomCli + ' ' + listaNom[h]
        h=h+1
    sheet.cell(row=i, column=7).value = NomCli    
    sheet.cell(row=i, column=8).value = apellido
    
    auxDirec=sheet.cell(row=i, column=14).value
    listaDir= [int(s) for s in auxDirec.split() if s.isdigit()]
    var01=len(listaDir)-1
    while var01>=0:
        if var01==0:
            codDirec=listaDir[var01]
            if codDirec == 0:
                codDirec=9999
            if codDirec == None:
                codDirec=9999
            break
        if len(str(listaDir[var01]))>1:
            codDirec=listaDir[var01]
            if codDirec == 0:
                codDirec=9999
            if codDirec == None:
                codDirec=9999
            break
             
        var01= var01-1
    sheet.cell(row=i, column=15).value=str(codDirec)
    # print('Nombre: ' + NomCli + '  Apellido: ' + apellido)
    # print('Direccion: ' + auxDirec + '  Numero: ' + str(codDirec) +'\n')
    
    i=i+1
workbook.save(r'X:\Logistica\1. Logistica Interno\Historiales_de_Logistica\ImportacionMasiva_Pruebas.xlsx') 
workbook  = openpyxl.load_workbook(r'X:\Logistica\1. Logistica Interno\Historiales_de_Logistica\ImportacionMasiva_Pruebas.xlsx')
sheet = workbook.active
CantArch = math.ceil((i-3)/regMaximos)
print('El Proceso finalizo.' + '\nSe procesaron ' + str(i-3) + ' registros con exito')
print('\nSe procesaron ' + str(CantArch) + ' archivos para subir a la Pagina de Andreani.')
print('Separando los registros en archivos de ' + str(regMaximos) + ' reg. como maximo')
print('Procesando.......')
# Copie el rango de celdas como una lista anidada
# Toma la celda inicial, la celda final y la hoja desde la que desea copiar.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)

    return rangeSelected
         

#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1
################################################
# Inicio de variables que usare a continuacion #
################################################        

regInicial=3
cio=1       #colInicialOrigen
fio=3       #filaInicialOrigen
cfo=29      #colFinalOrigen
ffo=regMaximos + 2     #filaFinalOrigen

cid=1       #colInicialDestino
fid=3       #filaInicialDestino
cfd=29      #colFinalDestino
ffd=regMaximos + 2     #filaFinalDestino
lista=[]
listaOp=[]
PopUp= 0
###############################################
# Creando coneccion a la pagina de Andreani   #
###############################################
options = webdriver.ChromeOptions()
# options.add_experimental_option('excludeSwitches', ['enable-logging'])
# options.add_argument('--headless')
browser = webdriver.Chrome(executable_path= r'C:\Users\eduardo.berga\AppData\Local\Programs\Python\Python38\chromedriver.exe', options=options)		
# browser = webdriver.Chrome(executable_path= r'C:\Users\seincomp\AppData\Local\anaconda3\chromedriver.exe', options=options)
browser.get('https://andreanionline.com/login')
browser.maximize_window()
try:
    element = WebDriverWait(browser, 200).until(EC.presence_of_element_located((By.ID,'main')))
    search_button = browser.find_element_by_id('loginButton')
    search_button.click()
except:
    print('Error al cargar el elemento >> loginButton <<\n')
    
try:
    element = WebDriverWait(browser, 200).until(EC.presence_of_element_located((By.ID,'username')))
    search_textinput = browser.find_element_by_id ('username')
    search_textinput.send_keys('xlshop@xl.com.ar')
    search_textinput = browser.find_element_by_id ('password')
    search_textinput.send_keys('Elarge,123')
    time.sleep(2)
    search_button = browser.find_element_by_xpath ('/html/body/div[2]/div/div[2]/form/div[4]/div[3]/button')
    search_button.click()
except:
    print('Error en la linea 157')

# Cargando archivo para pegar los registros de forma temporal
# Luego se borra el contenido
plantillaExcel = r'X:\Logistica\1. Logistica Interno\Historiales_de_Logistica\ImportacionAndreani.xlsx'

def vaciarCeldas(plantillaExcel,cio):
    template = openpyxl.load_workbook(plantillaExcel) #Add file name
    temp_sheet = template.active  #template["foo2"] #Add Sheet name
    ii=3
    while not(temp_sheet.cell(row=ii, column=1).value == None):
        while cio <= 29:
            if cio != 2:
                temp_sheet.cell(row=ii, column=cio).value = ''
                cio+=1
            else:
                cio+=1
        cio=1
        ii+=1
    template.save(plantillaExcel)    
   
if CantArch== 0:
    CantArch=1   
    
i=0    
while i < CantArch:
    # Dejar la plantilla en  blanco
    vaciarCeldas(plantillaExcel,cio)
    # Copiar registros a la plantilla
    template = openpyxl.load_workbook(plantillaExcel) #Add file name
    temp_sheet = template.active  #template["foo2"] #Add Sheet name
    fio = regInicial + (regMaximos * i)
    ffo = 202 + (regMaximos * i)
    selectedRange = copyRange(cio,fio,cfo,ffo,sheet)
    pastingRange = pasteRange(cid,fid,cfd,ffd,temp_sheet,selectedRange)
    template.save(plantillaExcel)
    template.close()
    
    fname = plantillaExcel
    #os.system ('taskkill / IM EXCEL.exe / F')
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = excel.Workbooks.Open(fname)
    #wb.DisplayAlerts = False
    wb.SaveAs(fname, FileFormat = 51)    #FileFormat = 51 is for .xlsx extension
    wb.Close()                               #FileFormat = 56 is for .xls extension
    # excel.Application.Quit()
    excel.DisplayAlerts = True
    # Carga plantilla en la pagina de Andreani
    # Retorna el numero de operacion
    browser.get('https://andreanionline.com/hacer-un-envio')
    time.sleep(10)
    try:
        element = WebDriverWait(browser, 200).until(EC.visibility_of_element_located((By.ID,'tabs-modo-de-carga')))
        time.sleep(2)
        search_button = browser.find_element_by_link_text('Carga masiva').click()
        element = WebDriverWait(browser, 200).until(EC.visibility_of_element_located((By.ID,'tabs-modo-de-carga')))
        time.sleep(3)
        #seleccionar lista desplegable de que enviar
        search_button = browser.find_element_by_xpath ('//*[@id="j_tipo_de_destino_y_envio_chosen"]').click()
        time.sleep(2)
        search_button = browser.find_element_by_xpath ('//*[@id="j_tipo_de_destino_y_envio_chosen"]/div/ul/li[2]').click()
        time.sleep(2)
        #seleccionar lista desplegable de tipo de envio
        search_button = browser.find_element_by_xpath ('//*[@id="j_modo_de_entrega_chosen"]').click()
        time.sleep(2)
        search_button = browser.find_element_by_xpath ('//*[@id="j_modo_de_entrega_chosen"]/div/ul/li[3]').click()
        time.sleep(5)         
        #seleccionar siguiente
        search_button=browser.find_element_by_id('btn-continuar-masiva-1')
        search_button.click()
        time.sleep(5)
    except:
         print('Error en la linea 220')
         
    try:
        element = WebDriverWait(browser, 200).until(EC.visibility_of_element_located((By.ID,'form-masiva-paso-2')))
        # Cargar archivo
        time.sleep(5)
        elem = browser.find_element_by_xpath("//input[@type='file']")
        elem.send_keys(plantillaExcel)
        time.sleep(10)
        #seleccionar boton siguiente
# =============================================================================
#         search_button=browser.find_element_by_id('btn-continuar-masiva-2')
#         search_button.click()
# =============================================================================
        print('El archivo se cargo con exito ')
    except:
        print('Error en la linea 240')
        
# =============================================================================
#     time.sleep(5)
#     # Selecciona sucursal de origen
#     print('Por seleccionar origen')
#     element = WebDriverWait(browser, 200).until(EC.visibility_of_element_located((By.ID,'btn-confirmar-masiva')))
#     search_button = browser.find_element_by_id('form-masiva-paso-3')
#     search_button = browser.find_element_by_xpath('/html/body/div[1]/div[1]/div[1]/div/div[2]/div[3]/div[2]/div[1]/form/div[1]/input').click()
#     search_button = browser.find_element_by_id('j_direccion_origen_domicilio_chosen').click()
#     time.sleep(5) 
#     
#     search_button = browser.find_element_by_xpath('/html/body/div[1]/div[1]/div[1]/div/div[2]/div[3]/div[2]/div[1]/form/div[2]/div[2]/div[1]/div/ul/li[2]').click()
# 
#     
#    
#     
#     # select_fr = Select(browser.find_element_by_id("j_direccion_origen_domicilio_chosen"))
#     # select_fr.select_by_index(1)
#     
#     
#     # select = Select(browser.find_element_by_id('j_direccion_origen_domicilio_chosen'))
# 
#     # # select by visible text
#     # select.select_by_visible_text('URUGUAY')
#     
#     # # select by value 
#     # select.select_by_value('1')
#     
#     # search_button.send_keys("{ENTER}")
#     # autoit.send("{ENTER}")
# 
# 
# 
# 
# #     search_button= browser.find_element_by_class_name('chosen-drop').click()
# #     search_button = browser.find_element_by_id('//*[@id="j_direccion_origen_domicilio_chosen"]/div/ul/li[2]')
# #     search_button = browser.find_element_by_id('//*[@id="j_direccion_origen_domicilio_chosen"]/div/ul/li[2]').send_keys("{ENTER}")
# #     search_button = browser.find_element_by_id("""CASA URUGUAY
# # 							(Uruguay 4415, ARGENTINA, BUENOS AIRES, VICTORIA, 1644)""").click()
# #     print(search_button)
#     # Presiona Confirmar
#     search_button=browser.find_element_by_id('btn-confirmar-masiva')
#     search_button.click()
#     time.sleep(10)
#     try:
#         element_present = EC.presence_of_element_located((By.ID,'j-checkbox-confirmacion'))
#         elemet= WebDriverWait(browser, 500).until(element_present)
#         # selecciona el check de responsabilidad
#         browser.execute_script("window.scrollBy(0, 50)") # Desplaza la pagina
#         search_button=browser.find_element_by_id('j-checkbox-confirmacion').click()
#         browser.execute_script("window.scrollBy(0, 180)") # Desplaza la pagina
#         # Desplaza la pagina al final
#         #browser.execute_script("window.scrollBy(0, window.innerHeight)") 
#          
#         time.sleep(5)
#         # selecciona el boton de confirmar
#         search_button=browser.find_element_by_partial_link_text('Confirmar')
#         search_button.click()
#         time.sleep(5)
#     except:
#         print('Error en el bloque de responsabilidad')
#         
#     try:
#         element = WebDriverWait(browser, 200).until(EC.visibility_of_element_located((By.CLASS_NAME,'message')))
#         elementos = browser.find_elements_by_xpath("//div[@class='message']/h4")
#         
#         for p in elementos:
#             valor = p.text
#             lista.append(valor)
#         
#         separador=" "
#         cadena=lista[i]
#         separado = cadena.split(separador)
#         # La variable SEPARADO[] contiene el numero de operacion
#        
#         listaOp.append(separado[3])
#     except:
#         print('Error en la linea xx')    
# =============================================================================
    i+=1
    
excel.Application.Quit()
time.sleep(5)
# =============================================================================
# ventana_actual= browser.window_handles[0]
# i=1
# print('Los Numero de Operacion generados son: \n') 
# print('>>>>>>>>>>>>>>>><<<<<<<<<<<<<<<')   
# for p in listaOp:
#     # Segundo bloque de automatizacion
#     # que se utilizara para la impresion de etiquetas
#     browser.get('https://andreanionline.com/envio')
#     time.sleep(5)
#     opnum= p
#     # opnum= '1836280'
#     try:
#         element = WebDriverWait(browser, 200).until(EC.presence_of_element_located((By.ID,'filtroGeneral')))
#         search_textinput = browser.find_element_by_id ('filtroGeneral')
#         search_textinput.send_keys(opnum)
#         time.sleep(5)
#         # Presiona el boton Buscar
#         search_textinput = browser.find_element_by_xpath('//*[@id="historial"]/div/div/div/form/div[1]/div[1]/button[1]')
#         search_textinput.click()
#         time.sleep(10)
#         # Mostrar 100 registros
#         search_button = browser.find_element_by_id('j-cantidad').click()
#         search_button = browser.find_element_by_xpath ('//*[@id="j-cantidad"]/option[4]').click()
#         time.sleep(10)
#         # Presiona el check seleccionar todos
#         browser.execute_script("window.scrollBy(0,-4000)") # Desplazamiento de la pagina hacia arriba
#         search_textinput = browser.find_element_by_xpath('//*[@id="historial"]/div/div/div/form/table/thead/tr/th[1]/input')
#         search_textinput.click()
#         # Presiona el boton Imprimir
#         time.sleep(3)
#         search_textinput = browser.find_element_by_id('imprimirConstanciasSeleccionadas')
#         search_textinput.click()
#         time.sleep(10)
#         # Guardar el valor de la nueva ventana
#         ventana_nueva=browser.window_handles[i]
#         # print(ventana_actual)
#         # print(ventana_nueva)
#         time.sleep(5)
#         browser.switch_to_window(ventana_nueva)
#     	# browser.maximize_window()
#         time.sleep(5)
#         autoit.send("^p")
#         time.sleep(15)
#         autoit.send("{ENTER}")
#         print('Se imprimio la operacion '+ p + '\n')
#         print('>>>>>>>>>>>>>>>><<<<<<<<<<<<<<<')
#         time.sleep(15)
#         browser.switch_to_window(ventana_actual)
#         i+=1
#     except:
#         print('Error en el bloque de impresion de etiquetas')
#         
# time.sleep(20)
# =============================================================================
browser.quit()
t_1=time.time()
t_t= t_1 - t_0
print('>>>>>>>>>>>>>>>><<<<<<<<<<<<<<<')
print('>>>>  El proceso finalizo  <<<<')
print('>>>>>>>>>>>>>>>><<<<<<<<<<<<<<<')
print('Tiempo total empleado ' + str(t_t))

# Volver a seleccionar la impresora por defecto
win32print.SetDefaultPrinter(currentprinter)
print('Se reestablecio la impresora predeterminada: \n'+ currentprinter)
for p in listaOp:
    print(p)
    len(p)

workbook  = openpyxl.load_workbook(r'X:\Logistica\1. Logistica Interno\Historiales_de_Logistica\ImportacionMasiva_Pruebas.xlsx')
workbook.save(r'C:\borrar\ImportacionMasiva_Pruebas.xlsx') 
