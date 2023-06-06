# -*- coding: utf-8 -*-
"""
Created on Mon Jun  1 11:44:50 2020

@author: eduardo.berga
"""





import openpyxl

# workbook  = openpyxl.load_workbook(r'C:\Users\eduardo.berga\Downloads\EjemploImportacionMasivaEncomiendaDomicilio.xlsx')
# sheet = workbook.active
# i=1
# encabezados=[]
# while not(sheet.cell(row=2, column=i).value == None):
#     encabezados.append(sheet.cell(row=2, column=i).value)
#     i= i+1


workbook  = openpyxl.load_workbook(r'X:\Logistica\1. Logistica Interno\Historiales_de_Logistica\ImportacionMasivaEncomiendaDomicilio.xlsx')
sheet = workbook.active

listaNom=[]
NomCli = str()
codDirec=0
i=3
while not(sheet.cell(row=i, column=1).value == None):
    aux=sheet.cell(row=i, column=7).value
    listaNom=aux.split(" ")
    aux01=len(listaNom)
    apellido=listaNom[aux01-1]
    h=0
    NomCli=''
    while h < aux01-1:
        NomCli= NomCli + ' ' + listaNom[h]
        h=h+1
    sheet.cell(row=i, column=7).value = NomCli    
    sheet.cell(row=i, column=8).value = apellido
    
    auxDirec=sheet.cell(row=i, column=14).value
    listaDir= [int(s) for s in auxDirec.split() if s.isdigit()]
    var01=len(listaDir)-1
    while var01>=0:
        if var01==0:
            codDirec=listaDir[var01]
            break
        if len(str(listaDir[var01]))>1:
            codDirec=listaDir[var01]
            break
             
        var01= var01-1
    sheet.cell(row=i, column=15).value=str(codDirec)
    print('Nombre: ' + NomCli + '  Apellido: ' + apellido)
    print('Direccion: ' + auxDirec + '  Numero: ' + str(codDirec) +'\n')
    
    i=i+1
workbook.save(r'X:\Logistica\1. Logistica Interno\Historiales_de_Logistica\ImportacionMasivaEncomiendaDomicilio.xlsx') 
print('El Proceso finalizo.' + '\nSe procesaron ' + i-1 + 'registros con exito')
