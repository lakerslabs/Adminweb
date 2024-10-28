import pprint
import os
from django.conf import settings
from django import template
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseRedirect,JsonResponse
from django.template import loader
from django.urls import reverse
from django.shortcuts import render, redirect,get_object_or_404
from Transportes.models import Transporte
from Transportes.forms import TransporteForm
from django.views import View
from django.views.generic.list import ListView
from apps.home.vistas.settingsUrls import *
from consultasTango.forms import TurnoForm,TurnoEditForm,CodigoErrorForm,ImageUploadForm
from consultasTango.models import Turno,CodigosError
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.core.files.storage import FileSystemStorage
import openpyxl
import pandas as pd
import json
from apps.home.SQL.Sql_Tango import validar_articulo,obtenerInformacionArticulo

# @login_required(login_url="/login/")
# def (request):
#     Nombre=''
#     dir_iframe = DIR_REPORTES['']
#     return render(request,'home/PlantillaHerramientas.html',{'dir_iframe':dir_iframe,'Nombre':Nombre})


# Logistica
@login_required(login_url="/login/")
def registro_turno(request):
    if request.method == 'POST':
        form = TurnoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('listado_turnos')  # Asumimos que crearemos esta vista más adelante
    else:
        form = TurnoForm()
    
    return render(request, 'appConsultasTango/registro_turno.html', {'form': form})

@login_required(login_url="/login/")
def get_nombre_proveedor(request):
    codigo_proveedor = request.GET.get('codigo', '')
    # Aquí deberías implementar la lógica para obtener el nombre del proveedor
    # Por ahora, usaremos un diccionario de ejemplo
    proveedores = {
        'AGODIF': 'DI FALCO MARIO DI FALCO JOSE Y DI FALCO COSME SOC DE HECHO',
        'BFBISE': 'BANCO MACRO S.A.',
        # ... Añade el resto de los proveedores aquí
    }
    nombre_proveedor = proveedores.get(codigo_proveedor, '')
    return JsonResponse({'nombre': nombre_proveedor})

def listado_turnos(request):
    turnos = Turno.objects.all().order_by('-FechaAsignacion')
    for turno in turnos:
        turno.progreso = calcular_progreso(turno)
        turno.save()

    # print(turnos)    
    return render(request, 'appConsultasTango/listado_turnos.html', {'turnos': turnos})

def eliminar_turno(request, turno_id):
    turno = get_object_or_404(Turno, pk=turno_id)
    if request.method == 'POST':
        turno.delete()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False})

def calcular_progreso(turno):
    total_steps = 3
    completed_steps = sum([turno.Recepcionado, turno.Auditado, turno.Posicionado])
    return int((completed_steps / total_steps) * 100)

def ver_turno(request, turno_id):
    turno = get_object_or_404(Turno, pk=turno_id)
    
    timeline = [
        {
            'estado': 'Recepcionado',
            'completado': turno.Recepcionado,
            'fecha': turno.RecepcionadoFechaHora
        },
        {
            'estado': 'Auditado',
            'completado': turno.Auditado,
            'fecha': turno.AuditadoFechaHora
        },
        {
            'estado': 'Posicionado',
            'completado': turno.Posicionado,
            'fecha': turno.PosicionadoFechaHora
        }
    ]
    
    context = {
        'turno': turno,
        'timeline': timeline
    }
    
    return render(request, 'appConsultasTango/ver_turno.html', context)

def editar_turno(request, turno_id):
    turno = get_object_or_404(Turno, pk=turno_id)
    if request.method == 'POST':
        form = TurnoEditForm(request.POST, instance=turno)
        if form.is_valid():
            turno = form.save(commit=False)
            
            # Actualizar las fechas de los estados
            if turno.Recepcionado and not turno.RecepcionadoFechaHora:
                turno.RecepcionadoFechaHora = timezone.now()
            if turno.Auditado and not turno.AuditadoFechaHora:
                turno.AuditadoFechaHora = timezone.now()
            if turno.Posicionado and not turno.PosicionadoFechaHora:
                turno.PosicionadoFechaHora = timezone.now()
            
            turno.save()
            return redirect('ver_turno', turno_id=turno.IdTurno)
    else:
        form = TurnoEditForm(instance=turno)
    
    return render(request, 'appConsultasTango/editar_turno.html', {'form': form, 'turno': turno})

def lista_codigos_error(request):
    codigos = CodigosError.objects.all().order_by('CodigoError')
    return render(request, 'appConsultasTango/lista_codigos_error.html', {'codigos': codigos})

def crear_codigo_error(request):
    if request.method == 'POST':
        form = CodigoErrorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Código de error creado exitosamente.')
            return redirect('lista_codigos_error')
    else:
        form = CodigoErrorForm()
    return render(request, 'appConsultasTango/crear_editar_codigo_error.html', {'form': form, 'accion': 'Crear'})

def editar_codigo_error(request, codigo_id):
    codigo = get_object_or_404(CodigosError, pk=codigo_id)
    if request.method == 'POST':
        form = CodigoErrorForm(request.POST, instance=codigo)
        if form.is_valid():
            form.save()
            messages.success(request, 'Código de error actualizado exitosamente.')
            return redirect('lista_codigos_error')
    else:
        form = CodigoErrorForm(instance=codigo)
    return render(request, 'appConsultasTango/crear_editar_codigo_error.html', {'form': form, 'accion': 'Editar'})

def eliminar_codigo_error(request, codigo_id):
    codigo = get_object_or_404(CodigosError, pk=codigo_id)
    if request.method == 'POST':
        codigo.delete()
        messages.success(request, 'Código de error eliminado exitosamente.')
        return redirect('lista_codigos_error')
    return render(request, 'appConsultasTango/confirmar_eliminar_codigo_error.html', {'codigo': codigo})
# @login_required(login_url="/login/")
# @method_decorator(csrf_exempt, name='dispatch')
class ImageUploadView(View):
    @method_decorator(login_required(login_url="/login/"))
    @method_decorator(csrf_exempt)
    def get(self, request):
        return render(request, 'appConsultasTango/uploadImg.html')

    def post(self, request):
        # Verificar si se recibieron archivos
        if not request.FILES:
            return JsonResponse({'error': 'No se recibieron archivos'}, status=400)

        files = []
        for key in request.FILES:
            files.extend(request.FILES.getlist(key))
        
        print('Files:', files)

        if not files:
            return JsonResponse({'error': 'No se recibieron archivos'}, status=400)

        uploaded_files = []
        for f in files:
            file_path = os.path.join(settings.MEDIA_ROOT, 'imgTempEcommerce', f.name)
            counter = 1
            # Asegurarse de que el nombre del archivo sea único
            while os.path.exists(file_path):
                name, extension = os.path.splitext(f.name)
                file_path = os.path.join(settings.MEDIA_ROOT, 'imgTempEcommerce', f"{name}_{counter}{extension}")
                counter += 1

            # Guardar el archivo en el sistema de archivos
            with open(file_path, 'wb+') as destination:
                for chunk in f.chunks():
                    destination.write(chunk)
            uploaded_files.append(f.name)

        return JsonResponse({'message': 'Imágenes cargadas con éxito', 'files': uploaded_files})

@login_required(login_url="/login/")
def upload_success(request):
    return render(request, 'appConsultasTango/successImg.html')

# ***************************
# ***************************

@login_required(login_url="/login/")
def Eliminar_Turno(request):
    if request.method == 'POST':
        IdTurno = request.POST.get('IdTurno')
        datos = Turno.objects.get(IdTurno=IdTurno)
        datos.delete()
        # Redirigir a la página de éxito
        return redirect('/../Herramientas/Calendario/TurnoListView')
    else:
        IdTurno = request.GET.get('IdTurno')
        datos = Turno.objects.get(IdTurno=IdTurno)
        return render(request, 'appConsultasTango/Eliminar_turno.html', {'datos': datos})

@login_required(login_url="/login/")
def Editar_Turno(request,IdTurno):
    datos = Turno.objects.get(IdTurno=IdTurno)
    if request.method == 'POST':
        form = TurnoForm(request.POST or None, request.FILES or None, instance=datos)
        if form.is_valid():
            form.save()
            # Redirigir a la página de éxito
            return redirect('/../Herramientas/Calendario/TurnoListView')
    else:
        form = TurnoForm(instance=datos)
    return render(request, 'appConsultasTango/Editar_turno.html', {'form': form})

@login_required(login_url="/login/")
def Listar_turno(request):
    datos = Turno.objects.all()   
    return render(request,'appConsultasTango/turno_list.html', {'turnos': datos})

@login_required(login_url="/login/")
def Crear_turno(request):
    if request.method == 'POST':
        form = TurnoForm(request.POST)
        if form.is_valid():
            form.save()
            # Redirigir a la página de éxito
            return redirect('/../Herramientas/Calendario/TurnoListView')
    else:
        form = TurnoForm()
    return render(request, 'appConsultasTango/Crear_turno.html', {'form': form})


@login_required(login_url="/login/")
def Gestion_cronograma(request):
    Nombre = 'Gestión cronograma'
    dir_iframe = DIR_HERAMIENTAS['Gestion_cronograma']
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe, })


@login_required(login_url="/login/")
def Gestion_guias_mayoristas(request):
    Nombre = 'Guías mayoristas'
    # dir_iframe = DIR_HERAMIENTAS['Gestion_guias_mayoristas']
    return render(request,'home/page-404.html')

# Abastecimiento

@login_required(login_url="/login/")
def Recodificacion(request):
    Nombre = 'Recodificacion Outlet'
    dir_iframe = DIR_HERAMIENTAS['Recodificacion']
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre})

@login_required(login_url="/login/")
def Stock_excluido(request):
    Nombre = 'Stock excluido'
    dir_iframe = DIR_HERAMIENTAS['Stock_excluido']
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe, })


@login_required(login_url="/login/")
def Carga_de_orden(request):
    Nombre = 'Carga de orden'
    dir_iframe = DIR_HERAMIENTAS['Carga_de_orden']
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe, })


@login_required(login_url="/login/")
def Activar_orden(request):
    Nombre = 'Activar orden'
    dir_iframe = DIR_HERAMIENTAS['Activar_orden']
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe, })


@login_required(login_url="/login/")
def Desactivar_orden(request):
    Nombre = 'Desactivar orden'
    dir_iframe = DIR_HERAMIENTAS['Desactivar_orden']
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe, })

@login_required(login_url="/login/")
def AltaPromoBancaria(request):
    Nombre = 'AltaPromoBancaria'
    dir_iframe = DIR_HERAMIENTAS['AltaPromoBancaria']
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe, })

@login_required(login_url="/login/")
def CrearGrupoPromo(request):
    Nombre = 'CrearGrupoPromo'
    dir_iframe = DIR_HERAMIENTAS['CrearGrupoPromo']
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe, })

@login_required(login_url="/login/")
def EditarGrupoPromo(request):
    Nombre = 'EditarGrupoPromo'
    dir_iframe = DIR_HERAMIENTAS['EditarGrupoPromo']
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe, })

# Comercial
@login_required(login_url="/login/")
def Ventas_sucursales(request):
    Nombre = 'Ventas sucursales'
    dir_iframe = DIR_HERAMIENTAS['Ventas_sucursales']
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe, })

@login_required(login_url="/login/")
def Gestion_categoria_productos(request):
    Nombre = ''
    dir_iframe = DIR_HERAMIENTAS['Gestion_categoria_productos'] #+ UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })

@login_required(login_url="/login/")
def AdministrarCuotas(request):
    Nombre = 'Administrar Cuotas'
    dir_iframe = DIR_HERAMIENTAS['AdministrarCuotas'] #+ UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })

@login_required(login_url="/login/")
def AdministrarInternos(request):
    Nombre = 'Administrar Internos'
    dir_iframe = DIR_HERAMIENTAS['AdministrarInternos'] #+ UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })

# Mayoristas
@login_required(login_url="/login/")
def Adm_Pedido(request):
    Nombre = 'Adm Pedido'
    dir_iframe = DIR_HERAMIENTAS['Adm_Pedido']
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe, })

# Ecommerce

@login_required(login_url="/login/")
def import_art_vtex(request):
    nombre_db='LAKER_SA'
    settings.DATABASES['mi_db_2']['NAME'] = nombre_db
    print('Cambiando base de datos a ' + nombre_db)
    print('import_art_vtex')
    try:
        if request.method == 'POST' and request.FILES['excel_file']:
            pfile = request.FILES['excel_file']
            filesys =FileSystemStorage()
            # Obtener el nombre del archivo sin espacios en blanco
            filename = pfile.name.replace(' ', '')

            # Guardar el archivo con el nombre sin espacios en blanco
            uploadfilename = filesys.save(filename, pfile)
            extension = os.path.splitext(uploadfilename)
            if  not extension[1] == '.xlsx':
                error_extension = 'El formato del archivo debe ser de tipo .xlsx'
                os.remove(filesys.path(uploadfilename))
                return render(request,'appConsultasTango/importFileArtVtex.html',{'mensaje_error':error_extension})

            uploaded_url = filesys.url(uploadfilename)  #Ruta donde se guardo el archivo
            uploaded_url = os.path.normpath(uploaded_url)
            # print("uploaded_url: " + uploaded_url)
            ruta_actual = os.path.join(os.getcwd()) #Ruta del proyecto
            # print('Ruta actual: ' + ruta_actual)
            path_filname = ruta_actual + uploaded_url
            wb = openpyxl.load_workbook(path_filname) #Abrimos el archivo
            worksheet = wb.active
            excel_data = list()
            enc_data = list()
            mensaje_error = ''
            mensaje_Success = ''
            art_valido = ''
            nombre_archivo = "AltaArtVtex.xls"
            eliminar_archivo_excel(nombre_archivo)
            # iterando sobre las filas y obteniendo
            # valor de cada celda en la fila
            for row in worksheet.iter_rows():
                fila = row[0].row   #Numero de fila
                row_data = list()
                talon_pedido = 0
                i = 1
                if worksheet.cell(row=fila, column=1).value is None: #Frena el loop al llegar al final de la lista
                    break
                
                for cell in row:
                    if fila == 1:
                        enc_data.append(str(cell.value))
                    else:
                        if worksheet.cell(row=1, column=i).value == 'ARTICULO':
                            if fila > 1:
                                articulo = worksheet.cell(row=fila, column=i).value
                                art_valido= validar_articulo(articulo)
                                print('art_valido: ' + str(art_valido))
                                if art_valido == 'ERROR':
                                    mensaje_error = 'Hay articulos que no existen en SJ_ETIQUETAS_FINAL'
                                    row_data.append('*' + str(cell.value) + '*')
                                    i += 1
                                    continue
                        elif worksheet.cell(row=1, column=i).value == 'DESCRIPCION':
                            worksheet.cell(row=fila, column=i).value = art_valido

                        if cell.value == None:
                            row_data.append(str(''))
                        else:
                            row_data.append(str(cell.value))                        
                    i += 1

                excel_data.append(row_data)
                if len(row_data) > 0:
                    resultado = json.loads(obtenerInformacionArticulo(row_data[0], row_data[2]))
                    crear_archivo_excel(resultado, nombre_archivo)
           
            wb.save(path_filname)
            if not(mensaje_error):
                mensaje_Success = 'Articulos cargados correctamente'
                os.remove(filesys.path(uploadfilename))                
            else:
                os.remove(filesys.path(uploadfilename))
            
            return render(request, 'appConsultasTango/importFileArtVtex.html' ,{'enc_data':enc_data,'excel_data':excel_data,'mensaje_Success':mensaje_Success,'mensaje_error':mensaje_error})


    except Exception as identifier:            
        print(identifier)
    return render(request,'appConsultasTango/importFileArtVtex.html',{})

import xlwt
import xlrd

def crear_archivo_excel(tempJson, nombre_archivo):
    # Construir la ruta absoluta del archivo
    ruta_archivo = os.path.join(settings.MEDIA_ROOT, nombre_archivo)

    try:
        # Abrir el archivo de Excel existente
        libro_rd = xlrd.open_workbook(ruta_archivo)
        hoja_rd = libro_rd.sheet_by_index(0)
        ultima_fila = hoja_rd.nrows

        # Crear un nuevo objeto Workbook
        libro_wt = xlwt.Workbook()
        hoja_wt = libro_wt.add_sheet('Hoja 1')

        # Copiar los datos de la hoja existente a la nueva hoja
        for i in range(ultima_fila):
            for j in range(hoja_rd.ncols):
                hoja_wt.write(i, j, hoja_rd.cell_value(i, j))

        # Agregar los nuevos datos a la hoja
        for i, fila in enumerate(tempJson):
            for j, valor in enumerate(fila.values()):
                hoja_wt.write(ultima_fila + i, j, valor)

        # Guardar el libro en un archivo
        libro_wt.save(ruta_archivo)
    except FileNotFoundError:
        # Si el archivo no existe, crear uno nuevo
        libro_wt = xlwt.Workbook()
        hoja_wt = libro_wt.add_sheet('Hoja 1')

        # Agregar los encabezados si es el primer registro
        if len(tempJson) > 0:
            encabezados = list(tempJson[0].keys())
            for i, encabezado in enumerate(encabezados):
                hoja_wt.write(0, i, encabezado)

        # Agregar los datos a la hoja
        for i, fila in enumerate(tempJson):
            for j, valor in enumerate(fila.values()):
                hoja_wt.write(i + 1, j, valor)

        # Guardar el libro en un archivo
        libro_wt.save(ruta_archivo)

def eliminar_archivo_excel(nombre_archivo):
    # Construir la ruta absoluta del archivo
    ruta_archivo = os.path.join(settings.MEDIA_ROOT, nombre_archivo)
    print('ruta_archivo: ' + ruta_archivo)
    # Verificar si el archivo existe y eliminarlo si es así
    if os.path.exists(ruta_archivo):
        os.remove(ruta_archivo)
        print('Se elimino el archivo ' + nombre_archivo)

@login_required(login_url="/login/")
def Control_pedidos(request):
    Nombre = 'Control pedidos'
    dir_iframe = DIR_HERAMIENTAS['Control_pedidos']
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe, })


@login_required(login_url="/login/")
def StockSegVtex(request):
    Nombre = 'Adm. Stock Seguridad Vtex'
    dir_iframe = DIR_HERAMIENTAS['StockSegVtex']
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe, })

@login_required(login_url="/login/")
def novICBC(request):
    Nombre = 'Actualizar novedades ICBC'
    dir_iframe = DIR_HERAMIENTAS['novICBC']
    # return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre})
    return redirect(dir_iframe)

# Gerencia
@login_required(login_url="/login/")
def rendircobranzas(request,UserName):
    Nombre = 'Rendir Cobranzas'
    dir_iframe = DIR_HERAMIENTAS['rendircobranzas'] + UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })

@login_required(login_url="/login/")
def GestionarCobro(request,UserName):
    Nombre = 'Gestionar Cobro'
    dir_iframe = DIR_HERAMIENTAS['gestionarCobro'] + UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })

@login_required(login_url="/login/")
def RegistrarEfectivo(request,UserName):
    Nombre = 'Registrar Efectivo'
    dir_iframe = DIR_HERAMIENTAS['registrarEfectivo'] + UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })

# Administracion

@login_required(login_url="/login/")
def CargaFacturasSuc(request):
    Nombre = 'Carga Facturas Suc'
    dir_iframe = DIR_HERAMIENTAS['CargaFacturasSuc'] #+ UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })
@login_required(login_url="/login/")
def CargaContratosFr(request):
    Nombre = 'Carga Contratos Franquicias'
    dir_iframe = DIR_HERAMIENTAS['CargaContratosFr'] #+ UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })

@login_required(login_url="/login/")
def ControlGastosSupervision(request):
    Nombre = 'Gastos Supervision'
    dir_iframe = DIR_HERAMIENTAS['ControlGastosSupervision'] #+ UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })

@login_required(login_url="/login/")
def ControlMasivoCobranza(request):
    Nombre = 'Control Gastos'
    dir_iframe = DIR_HERAMIENTAS['ControlMasivoCobranza'] #+ UserName
    return redirect(dir_iframe)

@login_required(login_url="/login/")
def Controlgastos(request):
    Nombre = 'Control Gastos'
    dir_iframe = DIR_HERAMIENTAS['controlGastos'] #+ UserName
    return redirect(dir_iframe)

@login_required(login_url="/login/")
def Cargargastos(request):
    Nombre = 'Gestionar Cobro'
    dir_iframe = DIR_HERAMIENTAS['cargaGastos'] #+ UserName
    return redirect(dir_iframe)

@login_required(login_url="/login/")
def Controlcajasdiario(request):
    Nombre = 'Control cajas Diario'
    dir_iframe = DIR_HERAMIENTAS['Controlcajasdiario'] #+ UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })

@login_required(login_url="/login/")
def CargaGastosAlquileres(request):
    Nombre = 'Carga Gastos Alquileres'
    dir_iframe = DIR_HERAMIENTAS['CargaGastosAlquileres'] #+ UserName
    return redirect(dir_iframe)

@login_required(login_url="/login/")
def GestionDeAlquileres(request):
    Nombre = 'Gestión % De Alquileres'
    dir_iframe = DIR_HERAMIENTAS['GestionDeAlquileres'] #+ UserName
    return redirect(dir_iframe)

@login_required(login_url="/login/")
def ControlEgresosDeCaja(request,UserName):
    Nombre = 'Control Egresos De Caja'
    dir_iframe = DIR_HERAMIENTAS['ControlEgresosDeCaja'] + UserName
    return redirect(dir_iframe)

@login_required(login_url="/login/")
def CargarContratosDeAlquiler(request):
    Nombre = 'Cargar Contratos De Alquiler'
    dir_iframe = DIR_HERAMIENTAS['CargarContratosDeAlquiler'] #+ UserName
    return redirect(dir_iframe)

@login_required(login_url="/login/")
def RelacionesCtaCont(request):
    Nombre = ''
    dir_iframe = DIR_HERAMIENTAS['RelacionesCtaCont'] #+ UserName
    return redirect(dir_iframe)

# Administracion_CE             ***Comercio Exterior***
@login_required(login_url="/login/")
def Cargarcontenedor(request):
    Nombre = 'Cargar Contenedor'
    dir_iframe = DIR_HERAMIENTAS['cargaInicial'] #+ UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })

@login_required(login_url="/login/")
def EditarContenedor(request):
    Nombre = 'Editar Contenedor'
    dir_iframe = DIR_HERAMIENTAS['mostrarOrden'] #+ UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })

# RRHH
@login_required(login_url="/login/")
def altaVendedores(request):
    Nombre = ''
    dir_iframe = DIR_HERAMIENTAS['altaVendedores'] #+ UserName
    # return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })
    return redirect(dir_iframe)

@login_required(login_url="/login/")
def listarGrupos(request):
    Nombre = ''
    dir_iframe = DIR_HERAMIENTAS['listarGrupos'] #+ UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })

@login_required(login_url="/login/")
def gestionarVendedores(request):
    Nombre = ''
    dir_iframe = DIR_HERAMIENTAS['gestionarVendedores'] #+ UserName
    # return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })
    return redirect(dir_iframe)

@login_required(login_url="/login/")
def adminEmpleados(request):
    Nombre = ''
    dir_iframe = DIR_HERAMIENTAS['adminEmpleados'] #+ UserName
    return render(request, 'home/PlantillaHerramientas.html', {'dir_iframe': dir_iframe,'Nombre':Nombre })
    # return redirect(dir_iframe)

